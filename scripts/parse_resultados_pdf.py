#!/usr/bin/env python3
"""
Parser de planillas de resultados en PDF — ADESCRUZ
Extrae resultados y genera JSON + Excel listos para cargar al sistema.

Uso: python3 parse_resultados_pdf.py <archivo.pdf> <concurso_id> <dia>
     dia: sab | dom
Ej:  python3 parse_resultados_pdf.py "SABADO Resumen IV CDS.pdf" "IV-CDS-2026" sab
"""

import sys, re, json
from pathlib import Path
import pdfplumber

# ── Mapeo categoría PDF → nombre oficial ─────────────────────────────────────
CAT_MAP = {
    'FUTUROS CAMPEONES':   'Futuros Campeones',
    'FUTUROS CAMPONES':    'Futuros Campeones',
    'ESCUELA MENOR':       'Escuela Menores',
    'ESCUELA MAYOR':       'Escuela Mayores',
    'ESCUELA MENORES':     'Escuela Menores',
    'ESCUELA MAYORES':     'Escuela Mayores',
    'PRE INFANTIL':        'Pre Infantil',
    'FOMENTO DEPORTIVO':   'Fomento Deportivo',
    'INFANTILES C':        'Infantil C',
    'INFANTIL C':          'Infantil C',
    'QUINTA CATEGORIA':    'Quinta Categoría',
    'QUINTA CATEGORÍA':    'Quinta Categoría',
    'CABALLOS NOVICIOS':   'Caballos Novicios',
    'INFANTIL B':          'Infantil B',
    'INFANTILES B':        'Infantil B',
    'CUARTA CATEGORIA':    'Cuarta Categoría',
    'CUARTA CATEGORÍA':    'Cuarta Categoría',
    'CABALLOS JOVENES S1': 'Caballos Jóvenes S1',
    'CABALLOS JÓVENES S1': 'Caballos Jóvenes S1',
    'INFANTIL A':          'Infantil A',
    'INFANTILES A':        'Infantil A',
    'TERCERA CATEGORIA':   'Tercera Categoría',
    'TERCERA CATEGORÍA':   'Tercera Categoría',
    'CABALLOS JOVENES S2': 'Caballos Jóvenes S2',
    'CABALLOS JÓVENES S2': 'Caballos Jóvenes S2',
    'PRE JUVENILES':       'Pre Juveniles',
    'SEGUNDA CATEGORIA':   'Segunda Categoría',
    'SEGUNDA CATEGORÍA':   'Segunda Categoría',
    'JUVENILES':           'Juveniles',
    'PRIMERA CATEGORIA':   'Primera Categoría',
    'PRIMERA CATEGORÍA':   'Primera Categoría',
    'ABIERTA':             'Abierta',
}

NO_RANKING       = {'Abierta'}
RANKING_CABALLO  = {'Caballos Novicios', 'Caballos Jóvenes S1', 'Caballos Jóvenes S2'}

# ── Puntos ────────────────────────────────────────────────────────────────────
def calcular_puntos(puesto, obs, total):
    obs_u   = str(obs or '').strip().upper()
    total_u = str(total or '').strip().upper()
    if obs_u in ('E', 'ELM', 'RET', 'NSP') or total_u in ('E', 'ELM', 'RET', 'NSP') or not obs_u:
        return 0
    try:
        if float(total_u.replace(',', '.')) > 12:
            return 0
    except ValueError:
        return 0
    try:
        p = int(str(puesto).strip())
    except (ValueError, TypeError):
        return 0
    return {1:7, 2:5, 3:4, 4:3, 5:2}.get(p, 1)


# ── Parser usando extract_tables() + posición en página ──────────────────────
def parse_pdf(pdf_path):
    resultados = []
    # Carry-state entre páginas: la última prueba/categoría vistas en la página
    # anterior. Necesario cuando el header "Prueba no. X" queda al final de una
    # página y sus tablas caen en la siguiente (caso IV CDS DOM: Prueba 3 y 5
    # perdidas sin este arrastre).
    carry = {'prueba': None, 'cat': None}

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_resultados, carry = parse_page(page, carry)
            resultados.extend(page_resultados)

    return resultados


def parse_page(page, carry_in=None):
    """Extrae resultados de una página combinando texto y tablas con coordenadas.
    carry_in: {'prueba': int|None, 'cat': str|None} — estado arrastrado desde
    la página anterior (última prueba/categoría vistas). Usado como fallback
    para tablas que aparecen antes de cualquier header nuevo en esta página.
    Retorna (resultados, carry_out)."""
    if carry_in is None:
        carry_in = {'prueba': None, 'cat': None}
    resultados = []

    # Extraer palabras con coordenadas (x0, y0, x1, y1, text)
    words = page.extract_words()

    # Extraer tablas con bounding boxes
    tables_bbox = page.find_tables()

    # Construir mapa de texto agrupado por línea (y0 redondeado)
    lines_by_y = {}
    for w in words:
        y = round(w['top'], 0)
        lines_by_y.setdefault(y, []).append(w)
    sorted_ys = sorted(lines_by_y.keys())

    # Reconstruir líneas de texto ordenadas por Y
    text_lines = []
    for y in sorted_ys:
        ws = sorted(lines_by_y[y], key=lambda w: w['x0'])
        text = ' '.join(w['text'] for w in ws)
        text_lines.append((y, text))

    # Detectar prueba actual desde el texto (antes de las tablas)
    prueba_y_map = {}   # y → prueba número (para saber qué prueba precede a cada tabla)

    for y, line in text_lines:
        m = re.match(r'^Prueba\s+no\.\s+(\d+)', line, re.IGNORECASE)
        if m:
            prueba_y_map[y] = int(m.group(1))

    # Detectar todos los headers de categoría en la página (y → cat oficial)
    cat_y_map = {}
    for y, line in text_lines:
        upper = line.strip().upper()
        if upper in CAT_MAP:
            cat_y_map[y] = CAT_MAP[upper]

    # Para cada tabla, encontrar:
    # 1. La prueba más cercana ANTES de la tabla (o carry_in.prueba)
    # 2. El nombre de categoría más cercano ANTES de la tabla (entre prueba y
    #    tabla, o carry_in.cat si no hay ni prueba ni cat nuevos antes)
    for tbl_obj in tables_bbox:
        bbox = tbl_obj.bbox  # (x0, top, x1, bottom)
        tbl_top = bbox[1]

        # Encontrar prueba más reciente antes de esta tabla en ESTA página
        prueba_antes = None
        prueba_y_antes = -1
        for y, pnum in prueba_y_map.items():
            if y < tbl_top and y > prueba_y_antes:
                prueba_y_antes = y
                prueba_antes = pnum

        # Fallback: si no hay prueba en esta página antes de la tabla, usar
        # la arrastrada desde la página anterior.
        if prueba_antes is None:
            prueba_antes = carry_in.get('prueba')
            # prueba_y_antes queda en -1 → cualquier cat en la página cuenta

        # Encontrar nombre de categoría más reciente antes de esta tabla (pero
        # después de la prueba más reciente en esta página)
        cat_antes = None
        cat_y_antes = prueba_y_antes
        for y, cname in cat_y_map.items():
            if y < tbl_top and y > cat_y_antes:
                cat_y_antes = y
                cat_antes = cname

        # Fallback: si no hay cat nueva en el rango y tampoco vino una prueba
        # nueva en la página (sea porque no hay, sea porque la tabla viene
        # antes del primer header), usar la cat arrastrada.
        if cat_antes is None and prueba_y_antes == -1:
            cat_antes = carry_in.get('cat')

        if prueba_antes is None or cat_antes is None:
            continue  # tabla sin contexto (ej: tabla de alturas)

        # Extraer filas de la tabla
        try:
            rows = tbl_obj.extract()
        except Exception:
            continue

        if not rows:
            continue

        # Saltar cabecera (fila con JINETE/AMAZONA)
        data_rows = [r for r in rows if r and r[0] and str(r[0]).strip().isdigit()]

        for row in data_rows:
            if len(row) < 6:
                continue

            num    = row[0]
            jinete = (row[1] or '').strip()
            caballo= (row[2] or '').strip()
            obs    = (row[3] or '').strip()
            tiempo = (row[4] or '').strip()
            ft     = (row[5] or '').strip() if len(row) > 5 else ''
            total  = (row[6] or '').strip() if len(row) > 6 else ''
            puesto = (row[7] or '').strip() if len(row) > 7 else ''

            if not jinete:
                continue

            puntos = calcular_puntos(puesto, obs, total)

            resultados.append({
                'prueba':           prueba_antes,
                'categoria':        cat_antes,
                'jinete':           jinete,
                'caballo':          caballo,
                'obs':              obs,
                'tiempo':           tiempo,
                'ft':               ft,
                'total':            total,
                'puesto':           puesto,
                'puntos':           puntos,
                'ranking':          cat_antes not in NO_RANKING,
                'ranking_caballo':  cat_antes in RANKING_CABALLO,
            })

    # Calcular carry_out: última prueba y cat vistas en esta página (por Y
    # mayor = más abajo), o las arrastradas si no hubo ninguna nueva.
    last_prueba = carry_in.get('prueba')
    if prueba_y_map:
        last_prueba = prueba_y_map[max(prueba_y_map.keys())]

    last_cat = carry_in.get('cat')
    if cat_y_map:
        last_cat = cat_y_map[max(cat_y_map.keys())]

    return resultados, {'prueba': last_prueba, 'cat': last_cat}


# ── Excel ─────────────────────────────────────────────────────────────────────
def exportar_excel(resultados, output_path, concurso_id, dia):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl no instalado — saltando Excel")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Resultados {dia.upper()}"

    headers = ['Concurso','Día','Prueba','Categoría','Puesto','Jinete','Caballo',
               'OBS','Tiempo','FT','Total Faltas','Puntos Ranking','Entra ranking','Ranking caballo']
    hfill = PatternFill(start_color='1a4731', end_color='1a4731', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for ri, r in enumerate(resultados, 2):
        row = [concurso_id, dia.upper(), r['prueba'], r['categoria'],
               r['puesto'], r['jinete'], r['caballo'], r['obs'],
               r['tiempo'], r['ft'], r['total'], r['puntos'],
               'Sí' if r['ranking'] else 'No',
               'Sí' if r['ranking_caballo'] else 'No']
        fill = PatternFill(start_color='f0f7f4' if ri%2==0 else 'FFFFFF',
                           end_color='f0f7f4' if ri%2==0 else 'FFFFFF', fill_type='solid')
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=ri, column=c, value=v)
            cell.fill = fill

    for c, w in enumerate([14,6,8,22,8,28,22,8,10,8,14,16,14,18], 1):
        ws.column_dimensions[ws.cell(row=1,column=c).column_letter].width = w

    wb.save(output_path)
    print(f"Excel guardado: {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 4:
        print(__doc__); sys.exit(1)

    pdf_path    = sys.argv[1]
    concurso_id = sys.argv[2]
    dia         = sys.argv[3].lower()

    print(f"Procesando: {pdf_path}")
    resultados = parse_pdf(pdf_path)

    # Normalizar status codes a forma canonical uppercase (convención ADESCRUZ)
    STATUS_MAP = {'E': 'ELM', 'ELM': 'ELM', 'RET': 'RET', 'NSP': 'NSP'}
    for r in resultados:
        for k in ('obs', 'tiempo', 'ft', 'total'):
            v = str(r.get(k) or '').strip().upper()
            if v in STATUS_MAP:
                r[k] = STATUS_MAP[v]

    print(f"✓ {len(resultados)} resultados extraídos\n")

    from collections import Counter
    for cat, n in sorted(Counter(r['categoria'] for r in resultados).items()):
        print(f"  {cat}: {n}")

    base = Path(pdf_path).parent / Path(pdf_path).stem
    json_path = Path(str(base) + '_parsed.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump({'concurso_id': concurso_id, 'dia': dia,
                   'total': len(resultados), 'resultados': resultados},
                  f, ensure_ascii=False, indent=2)
    print(f"\nJSON: {json_path}")

    xlsx_path = Path(str(base) + '_parsed.xlsx')
    exportar_excel(resultados, xlsx_path, concurso_id, dia)
    return resultados

if __name__ == '__main__':
    main()
